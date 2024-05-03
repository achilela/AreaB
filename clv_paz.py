import streamlit as st
import pandas as pd
import time
from io import BytesIO
from xlrd import open_workbook
from xlsxwriter import Workbook
from datetime import date

# Set page title
st.set_page_config(page_title="Topsides Plant Maintenance Data Analysis")

# Add a title and description
st.markdown(
    """
    <h1 style='text-align: center; font-size: 36px; color: #2F80ED;'>
        Topsides Plant Maintenance Data Analysis
    </h1>
    
    <p style='text-align: center; font-size: 18px;'>
        Upload an Excel file and use the dropdown menus to filter and analyze the data.
    </p>
    """,
    unsafe_allow_html=True
)

# File upload in the sidebar
uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Show progress bar during file upload
    progress_text = "Uploading file..."
    my_bar = st.progress(0, text=progress_text)

    for percent_complete in range(100):
        time.sleep(0.01)
        my_bar.progress(percent_complete + 1, text=progress_text)

    # Load the Excel file into a BytesIO object
    excel_data = BytesIO(uploaded_file.getvalue())

    # Load the workbook and select the "Data Base" sheet
    workbook = open_workbook(file_contents=excel_data.getvalue())
    sheet = workbook.sheet_by_name("Data Base")

    # Delete the first column (A) and the last three columns (S, T, U)
    data = [sheet.row_values(r) for r in range(sheet.nrows)]
    data = [[row[1]] + row[4:-3] for row in data]

    # Delete the first 4 rows
    data = data[4:]

    # Add a new column with the header "Today's Date" and insert the TODAY() formula
    data[0].append("Today's Date")
    for row in data[1:]:
        row.append(date.today())

    # Create a new workbook and sheet using xlsxwriter
    output = BytesIO()
    new_workbook = Workbook(output, {'in_memory': True})
    new_sheet = new_workbook.add_worksheet()

    # Write the data to the new sheet
    for r, row in enumerate(data):
        for c, value in enumerate(row):
            new_sheet.write(r, c, value)

    # Save the new workbook
    new_workbook.close()
    output.seek(0)

    # Read the modified Excel file into a pandas DataFrame
    df = pd.read_excel(output)

    # Check if the Excel file is already in table form
    if df.columns.nlevels == 1:
        # Remove rows with missing data
        df.dropna(how='all', inplace=True)

        # Remove columns with missing data
        df.dropna(axis=1, how='all', inplace=True)

        # Reset index
        df.reset_index(drop=True, inplace=True)

        # Replace NaN values in the "SECE STATUS" column with "Non-SCE"
        if "SECE STATUS" in df.columns:
            df["SECE STATUS"].fillna("Non-SCE", inplace=True)

        # Display the first 10 rows of the cleaned table
        st.write("Cleaned Table:")
        st.write(df.head(10))

        # Get unique values for dropdown menus
        column_options = df.columns.tolist()

        # Create radio button to select the main column
        main_column = st.radio("Select the main column", column_options)

        # Create multiselect dropdowns for filtering other columns
        filter_columns = [col for col in column_options if col != main_column]
        selected_columns = st.multiselect("Select columns to filter", filter_columns)

        if selected_columns:
            # Filter the DataFrame based on selected columns
            filtered_df = df[[main_column] + selected_columns]

            # Create a dictionary to store the filter values for each selected column
            filter_values = {}

            # Create multiselect dropdowns for filtering each selected column
            for column in selected_columns:
                unique_values = filtered_df[column].unique()
                filter_values[column] = st.multiselect(f"Select values to filter '{column}'", unique_values)

            # Filter the DataFrame based on the filter values
            for column, values in filter_values.items():
                if values:
                    filtered_df = filtered_df[filtered_df[column].isin(values)]

            # Group the data by the main column and selected columns
            grouped_data = filtered_df.groupby([main_column] + selected_columns).size().reset_index(name='Count')

            # Pivot the grouped data to create a table with the main column as rows and selected columns as columns
            pivot_table = grouped_data.pivot_table(index=main_column, columns=selected_columns, values='Count', fill_value=0)

            # Add a "Grand Total" column to the pivot table
            pivot_table["Grand Total"] = pivot_table.sum(axis=1)

            # Add a "Total" row to the pivot table
            pivot_table.loc["Total"] = pivot_table.sum()

            # Display the pivot table
            st.write("Filtered Table:")
            st.write(pivot_table)

    else:
        st.write("The uploaded Excel file is not in table form.")
In this updated code, we use the xlrd package to read the Excel file and the xlsxwriter package to write the modified data to a new Excel file. Here's a summary of the changes:

We import the xlrd and xlsxwriter packages instead of openpyxl.
We use open_workbook() from xlrd to load the workbook and sheet_by_name() to select the "Data Base" sheet.
We delete the first column (A) and the last three columns (S, T, U) by slicing the row values.
We delete the first 4 rows by slicing the data list.
We add a new column with the header "Today's Date" and insert the date.today() value for each row.
We create a new workbook and sheet using xlsxwriter and write the data to the new sheet.
We save the new workbook to a BytesIO object.
The rest of the code remains the same, reading the modified Excel file into a pandas DataFrame and performing the data analysis.

Make sure to install the xlrd and xlsxwriter packages by running the following commands:
Copy codepip install xlrd
pip install xlsxwriter
This alternative code using xlrd and xlsxwriter should work as a replacement for the code using openpyxl. If you still encounter any issues, please let me know, and I'll be happy to assist you further. CopyClaude does not have the ability to run the code it generates yet.AMYour task as a methods engineer data analyst is to consider the two slides attached one as week17 and other as week18. You will compare all grandtotals and totals and performance providing comparison responses in terms of percentage increase and KPIs. Be succinct and to the point.Comparing the grand totals and totals between Week 17 and Week 18, here are the key observations:

Grand Total Progress:

Week 17: 50
Week 18: 44
The grand total progress decreased by 12% from Week 17 to Week 18.


Total Progress (based on "Total" row):

Week 17: 59
Week 18: 46
The total progress decreased by approximately 22% from Week 17 to Week 18.


Pressure Safety Device:

Week 17: 15
Week 18: 12
The count of Pressure Safety Devices decreased by 20% from Week 17 to Week 18.


Pressure Vessel (VIE):

Week 17: 9
Week 18: 10
The count of Pressure Vessels (VIE) increased by approximately 11% from Week 17 to Week 18.


Piping:

Week 17: 17
Week 18: 12
The count of Piping decreased by approximately 29% from Week 17 to Week 18.


Corrosion Monitoring:

Week 17: 4
Week 18: 4
The count of Corrosion Monitoring remained the same between Week 17 and Week 18.


VIE + NOT in lieu of VIE & removed from 2024:

Week 17: 61
Week 18: 61
The count of VIE + NOT remained the same between Week 17 and Week 18.


PIPING Progress:

Week 17: 47%
Week 18: 39%
The PIPING progress percentage decreased by 8 percentage points from Week 17 to Week 18.


Site Perf Progress:

Week 17: 82%
Week 18: 95%
The Site Perf progress percentage increased by 13 percentage points from Week 17 to Week 18.


Plan Progress:

Week 17: 66%
Week 18: 69%
The Plan progress percentage increased by 3 percentage points from Week 17 to Week 18.



In summary, while the grand total and total progress decreased from Week 17 to Week 18, there were some improvements in specific areas such as Pressure Vessel (VIE) count and Site Perf progress percentage. However, the overall progress seems to have declined, particularly in Pressure Safety Device, Piping, and PIPING progress percentage. Copyclv_paztext/x-pythonAMThe code above needs to provide functionalities to handle open and manipulate or disable all macros within the .xlsm file. Provide code to open and handle excel files with macro. Keep your code structure the same, however, just add python functionalities to handle excel with macro vbaTo handle Excel files with macros (.xlsm files) and provide functionalities to manipulate or disable macros, you can use the openpyxl library along with the win32com module. Here's the updated code with the added functionality to handle Excel files with macros:
pythonCopy codeimport streamlit as st
import pandas as pd
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date
import dill as pickle
import win32com.client

# Set page title
st.set_page_config(page_title="Topsides Plant Maintenance Data Analysis")

# Add a title and description
st.markdown(
    """
    <h1 style='text-align: center; font-size: 36px; color: #2F80ED;'>
        Topsides Plant Maintenance Data Analysis
    </h1>

    <p style='text-align: center; font-size: 18px;'>
        Upload an Excel file and use the dropdown menus to filter and analyze the data.
    </p>
    """,
    unsafe_allow_html=True
)

# File upload in the sidebar
uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx", "xls", "xlsm"])

@st.cache_resource
def load_data(uploaded_file):
    if uploaded_file is not None:
        # Show progress bar during file upload
        progress_text = "Uploading file..."
        my_bar = st.progress(0, text=progress_text)

        for percent_complete in range(100):
            time.sleep(0.01)
            my_bar.progress(percent_complete + 1, text=progress_text)

        # Load the Excel file into a BytesIO object
        excel_data = BytesIO(uploaded_file.getvalue())

        # Check if the file has macros
        if uploaded_file.type == "application/vnd.ms-excel.sheet.macroEnabled.12":
            # Open the workbook with macros using win32com
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(excel_data)

            # Disable macros
            excel.EnableEvents = False
            excel.DisplayAlerts = False

            # Save the workbook without macros
            output = BytesIO()
            workbook.SaveAs(output, FileFormat=51)  # 51 represents the .xlsx format
            output.seek(0)

            # Close the workbook and quit Excel
            workbook.Close()
            excel.Quit()

            # Load the workbook from the BytesIO object
            workbook = load_workbook(output)
        else:
            # Load the workbook directly from the BytesIO object
            workbook = load_workbook(excel_data)

        # Select the "Data Base" sheet
        sheet = workbook["Data Base"]

        # Delete the first column (A) and the last three columns (S, T, U)
        sheet.delete_cols(1, 1)
        sheet.delete_cols(sheet.max_column - 2, 3)

        # Delete the first 4 rows
        sheet.delete_rows(1, 4)

        # Add a new column with the header "Today's Date" and insert the TODAY() formula
        sheet.cell(row=1, column=sheet.max_column + 1, value="Today's Date")
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=sheet.max_column, value=f"=TODAY()")

        # Convert the sheet to a table
        table_name = "MainTable"
        table_range = f"A1:{chr(ord('A') + sheet.max_column - 1)}{sheet.max_row}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        # Save the modified workbook to a BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Read the modified Excel file into a pandas DataFrame
        df = pd.read_excel(output)

        # Check if the Excel file is already in table form
        if df.columns.nlevels == 1:
            # Remove rows with missing data
            df.dropna(how='all', inplace=True)

            # Remove columns with missing data
            df.dropna(axis=1, how='all', inplace=True)

            # Reset index
            df.reset_index(drop=True, inplace=True)

            # Replace NaN values in the "SECE STATUS" column with "Non-SCE"
            if "SECE STATUS" in df.columns:
                df["SECE STATUS"].fillna("Non-SCE", inplace=True)

        return df

df = load_data(uploaded_file)

if df is not None:
    # Display the first 10 rows of the cleaned table
    st.write("Cleaned Table:")
    st.write(df.head(10))

    # Get unique values for dropdown menus
    column_options = df.columns.tolist()

    # Create radio button to select the main column
    main_column = st.radio("Select the main column", column_options)

    # Create multiselect dropdowns for filtering other columns
    filter_columns = [col for col in column_options if col != main_column]
    selected_columns = st.multiselect("Select columns to filter", filter_columns)

    if selected_columns:
        # Filter the DataFrame based on selected columns
        filtered_df = df[df[main_column].isin(values)]

        # Create a dictionary to store the filter values for each selected column
        filter_values = {}

        # Create multiselect dropdowns for filtering each selected column
        for column in selected_columns:
            unique_values = filtered_df[column].unique()
            filter_values[column] = st.multiselect(f"Select values to filter '{column}'", unique_values)

        # Filter the DataFrame based on the filter values
        for column, values in filter_values.items():
            if values:
                filtered_df = filtered_df[filtered_df[column].isin(values)]

        # Group the data by the main column and selected columns
        grouped_data = filtered_df.groupby([main_column] + selected_columns).size().reset_index(name='Count')

        # Pivot the grouped data to create a table with the main column as rows and selected columns as columns
        pivot_table = grouped_data.pivot_table(index=main_column, columns=selected_columns, values='Count', fill_value=0)

        # Add a "Grand Total" column to the pivot table
        pivot_table["Grand Total"] = pivot_table.sum(axis=1)

        # Add a "Total" row to the pivot table
        pivot_table.loc["Total"] = pivot_table.sum()

        # Display the pivot table
        st.write("Filtered Table:")
        st.write(pivot_table)

    else:
        st.write("The uploaded Excel file is not in table form.")
The main changes in the code are:

The win32com module is imported to handle Excel files with macros.
In the load_data function, a check is added to determine if the uploaded file has macros by checking the file type (uploaded_file.type). If it is a macro-enabled file (.xlsm), the following steps are performed:

The workbook is opened using win32com.client.Dispatch("Excel.Application").
Macros are disabled by setting excel.EnableEvents = False and excel.DisplayAlerts = False.
The workbook is saved without macros to a new BytesIO object using workbook.SaveAs(output, FileFormat=51), where 51 represents the .xlsx format.
The workbook is closed, and Excel is quit using workbook.Close() and excel.Quit().
The modified workbook is loaded from the BytesIO object using load_workbook(output).


If the uploaded file is not a macro-enabled file, the workbook is loaded directly from the BytesIO object using load_workbook(excel_data).
The rest of the code remains the same, with the added functionality to handle Excel files with macros.

Note: The win32com module is a Windows-specific module and requires the pywin32 package to be installed. You can install it by running pip install pywin32 in your command prompt or terminal.
With these modifications, the code can now handle both regular Excel files (.xlsx and .xls) and Excel files with macros (.xlsm). The macros will be disabled, and the file will be saved as a regular Excel file before processing. CopyClaude does not have the ability to run the code it generates yet.clv_paztext/x-pythonAMThe code above needs to provide functionalities to handle open and manipulate or disable all macros within the .xlsm file. Provide code to open and handle excel files with macro. Keep your code structure the same, however, just add python functionalities to handle excel with macro vbaTo handle Excel files with macros (.xlsm files) and provide functionalities to manipulate or disable macros, you can use the openpyxl library along with the win32com module. Here's the updated code with the added functionality to handle Excel files with macros:
pythonCopy codeimport streamlit as st
import pandas as pd
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date
import dill as pickle
import win32com.client

# Set page title
st.set_page_config(page_title="Topsides Plant Maintenance Data Analysis")

# Add a title and description
st.markdown(
    """
    <h1 style='text-align: center; font-size: 36px; color: #2F80ED;'>
        Topsides Plant Maintenance Data Analysis
    </h1>

    <p style='text-align: center; font-size: 18px;'>
        Upload an Excel file and use the dropdown menus to filter and analyze the data.
    </p>
    """,
    unsafe_allow_html=True
)

# File upload in the sidebar
uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx", "xls", "xlsm"])

@st.cache_resource
def load_data(uploaded_file):
    if uploaded_file is not None:
        # Show progress bar during file upload
        progress_text = "Uploading file..."
        my_bar = st.progress(0, text=progress_text)

        for percent_complete in range(100):
            time.sleep(0.01)
            my_bar.progress(percent_complete + 1, text=progress_text)

        # Load the Excel file into a BytesIO object
        excel_data = BytesIO(uploaded_file.getvalue())

        # Check if the file has macros
        if uploaded_file.type == "application/vnd.ms-excel.sheet.macroEnabled.12":
            # Open the workbook with macros using win32com
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(excel_data)

            # Disable macros
            excel.EnableEvents = False
            excel.DisplayAlerts = False

            # Save the workbook without macros
            output = BytesIO()
            workbook.SaveAs(output, FileFormat=51)  # 51 represents the .xlsx format
            output.seek(0)

            # Close the workbook and quit Excel
            workbook.Close()
            excel.Quit()

            # Load the workbook from the BytesIO object
            workbook = load_workbook(output)
        else:
            # Load the workbook directly from the BytesIO object
            workbook = load_workbook(excel_data)

        # Select the "Data Base" sheet
        sheet = workbook["Data Base"]

        # Delete the first column (A) and the last three columns (S, T, U)
        sheet.delete_cols(1, 1)
        sheet.delete_cols(sheet.max_column - 2, 3)

        # Delete the first 4 rows
        sheet.delete_rows(1, 4)

        # Add a new column with the header "Today's Date" and insert the TODAY() formula
        sheet.cell(row=1, column=sheet.max_column + 1, value="Today's Date")
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=sheet.max_column, value=f"=TODAY()")

        # Convert the sheet to a table
        table_name = "MainTable"
        table_range = f"A1:{chr(ord('A') + sheet.max_column - 1)}{sheet.max_row}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        # Save the modified workbook to a BytesIO object
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Read the modified Excel file into a pandas DataFrame
        df = pd.read_excel(output)

        # Check if the Excel file is already in table form
        if df.columns.nlevels == 1:
            # Remove rows with missing data
            df.dropna(how='all', inplace=True)

            # Remove columns with missing data
            df.dropna(axis=1, how='all', inplace=True)

            # Reset index
            df.reset_index(drop=True, inplace=True)

            # Replace NaN values in the "SECE STATUS" column with "Non-SCE"
            if "SECE STATUS" in df.columns:
                df["SECE STATUS"].fillna("Non-SCE", inplace=True)

        return df

df = load_data(uploaded_file)

if df is not None:
    # Display the first 10 rows of the cleaned table
    st.write("Cleaned Table:")
    st.write(df.head(10))

    # Get unique values for dropdown menus
    column_options = df.columns.tolist()

    # Create radio button to select the main column
    main_column = st.radio("Select the main column", column_options)

    # Create multiselect dropdowns for filtering other columns
    filter_columns = [col for col in column_options if col != main_column]
    selected_columns = st.multiselect("Select columns to filter", filter_columns)

    if selected_columns:
        # Filter the DataFrame based on selected columns
        filtered_df = df[df[main_column].isin(values)]

        # Create a dictionary to store the filter values for each selected column
        filter_values = {}

        # Create multiselect dropdowns for filtering each selected column
        for column in selected_columns:
            unique_values = filtered_df[column].unique()
            filter_values[column] = st.multiselect(f"Select values to filter '{column}'", unique_values)

        # Filter the DataFrame based on the filter values
        for column, values in filter_values.items():
            if values:
                filtered_df = filtered_df[filtered_df[column].isin(values)]

        # Group the data by the main column and selected columns
        grouped_data = filtered_df.groupby([main_column] + selected_columns).size().reset_index(name='Count')

        # Pivot the grouped data to create a table with the main column as rows and selected columns as columns
        pivot_table = grouped_data.pivot_table(index=main_column, columns=selected_columns, values='Count', fill_value=0)

        # Add a "Grand Total" column to the pivot table
        pivot_table["Grand Total"] = pivot_table.sum(axis=1)

        # Add a "Total" row to the pivot table
        pivot_table.loc["Total"] = pivot_table.sum()

        # Display the pivot table
        st.write("Filtered Table:")
        st.write(pivot_table)

    else:
        st.write("The uploaded Excel file is not in table form.")
